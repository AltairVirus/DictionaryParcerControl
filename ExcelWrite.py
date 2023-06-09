import openpyxl
import Config
import datetime


def read_massive_from_excel(file_name=Config.data_excel,
                            min_row=Config.min_col,
                            max_row=Config.max_col,
                            min_col=Config.min_row,
                            max_col=Config.max_row):

    book = openpyxl.load_workbook(file_name)
    sheet = book.worksheets[0]

    # print(tuple(sheet.rows))
    reference_uniq_cols = []

    for col in sheet.iter_cols(min_row, max_row, min_col, max_col):
        help_list = []
        for cell in col:
            if cell.value not in help_list:
                help_list.append(cell.value)
        reference_uniq_cols.append(help_list)

    return reference_uniq_cols


def write_massive_to_excel(massive, col=1, row=1):

    wb = openpyxl.Workbook()
    ws = wb.active

    if type(massive) == list:
        for subarray in massive:
            for index, value in enumerate(subarray):
                ws.cell(column=col + index, row=row).value = value
            row += 1
    elif type(massive) == dict:
        for subarray in massive:
            ws.cell(column=col, row=row).value = subarray
            for index, value in enumerate(massive.get(subarray)):
                ws.cell(column=col, row=row + 1 + index).value = value
            col += 1

    dt = datetime.datetime.now()
    time = dt.time()
    now = str(time.strftime('%H ч %M мин'))

    wb.save(f"./{now} " + ("Атрибуты" if type(massive) == dict else "Значения") + ".xlsx")